from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Optional

_ALLOWED_SCORE_FIELDS = frozenset({"ct_score1", "ct_score2", "ct_score3", "lab_score"})


@dataclass
class ParsedMarkRow:
    row: int
    roll_number: str
    ct_score1: Optional[float] = None
    ct_score2: Optional[float] = None
    ct_score3: Optional[float] = None
    lab_score: Optional[float] = None


@dataclass
class ParseError:
    row: int
    roll_number: Optional[str] = None
    reason: str = ""


@dataclass
class ParseResult:
    records: list[ParsedMarkRow] = field(default_factory=list)
    errors: list[ParseError] = field(default_factory=list)


_ROLL_NAMES = {
    "rollnumber",
    "roll_number",
    "role_number",  # e.g. "Role Number"
    "rolenumber",
    "roll",
    "studentid",
    "student_id",
    "id",
}
_SCORE_MAP: dict[str, str] = {
    "ct1": "ct_score1", "ctscore1": "ct_score1", "ct_score_1": "ct_score1",
    "ct2": "ct_score2", "ctscore2": "ct_score2", "ct_score_2": "ct_score2",
    "ct3": "ct_score3", "ctscore3": "ct_score3", "ct_score_3": "ct_score3",
    "lab": "lab_score", "labscore": "lab_score", "lab_score": "lab_score",
}

_GENERIC_SCORE_KEYS = ("mark", "marks", "score", "total", "obtained", "value", "ct")


def _norm(h: str) -> str:
    """Normalize header labels for matching (slashes as in ``Marks / 20`` → ``marks_20``)."""
    return re.sub(r"[\s\-/]+", "_", h.strip().lower())


def _append_marks_out_of_columns(header_map: dict[str, str], score_mapping: list[tuple[str, str]]) -> None:
    """Map ``Marks / N`` style columns to ct_score1..3 by descending N (first / highest max → CT1)."""
    used_targets = {t for _, t in score_mapping}
    candidates: list[tuple[int, str, str]] = []
    for nk, oh in header_map.items():
        m = re.match(r"^marks_(\d+)$", nk)
        if m:
            candidates.append((int(m.group(1)), nk, oh))
    candidates.sort(key=lambda x: -x[0])
    fields = ["ct_score1", "ct_score2", "ct_score3"]
    fi = 0
    for _pts, _nk, oh in candidates:
        while fi < len(fields) and fields[fi] in used_targets:
            fi += 1
        if fi >= len(fields):
            break
        tgt = fields[fi]
        score_mapping.append((oh, tgt))
        used_targets.add(tgt)
        fi += 1


def infer_assessment_field_from_label(label: str | None) -> str | None:
    """Map tutor label (e.g. 'CT 2', 'Class Test 3') to Enrollment column attribute name."""
    if not label or not str(label).strip():
        return None
    compact = re.sub(r"[\s_\-]+", "", str(label).strip().lower())
    if "lab" in compact:
        return "lab_score"
    if re.search(r"ct[^0-9]*3|classtest[^0-9]*3|^class3$", compact):
        return "ct_score3"
    if re.search(r"ct[^0-9]*2|classtest[^0-9]*2|^class2$", compact):
        return "ct_score2"
    if re.search(r"ct[^0-9]*1|classtest[^0-9]*1|^class1$", compact):
        return "ct_score1"
    return None


def _normalize_roll(value: str) -> str:
    return " ".join(str(value).split())


def _load_tabular_rows(data: bytes, filename: str) -> list[dict] | ParseError:
    lower = filename.lower()
    try:
        if lower.endswith(".csv"):
            text = data.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            return list(reader)
        if lower.endswith(".xlsx"):
            import openpyxl  # type: ignore

            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            if ws is None:
                return ParseError(row=0, reason="Spreadsheet has no sheets")
            iter_rows = list(ws.iter_rows(values_only=True))
            if not iter_rows:
                return ParseError(row=0, reason="Spreadsheet is empty")
            headers = [str(c) if c is not None else "" for c in iter_rows[0]]
            return [
                dict(zip(headers, [str(c) if c is not None else "" for c in row]))
                for row in iter_rows[1:]
            ]
        return ParseError(row=0, reason="Unsupported format (use .csv or .xlsx)")
    except Exception as exc:
        return ParseError(row=0, reason=f"Failed to parse file: {exc}")


def _parse_rows(rows: list[dict], assessment_field: str | None = None) -> ParseResult:
    if not rows:
        return ParseResult(errors=[ParseError(row=0, reason="Spreadsheet is empty")])

    if assessment_field is not None and assessment_field not in _ALLOWED_SCORE_FIELDS:
        return ParseResult(errors=[ParseError(row=0, reason="Invalid assessment field mapping")])

    headers = list(rows[0].keys())
    header_map = {_norm(h): h for h in headers}

    roll_key = next((k for k in _ROLL_NAMES if k in header_map), None)
    if not roll_key:
        return ParseResult(
            errors=[ParseError(row=0, reason="No roll number column found. Expected: rollNumber, roll, studentId")]
        )
    roll_header = header_map[roll_key]

    score_mapping: list[tuple[str, str]] = [(header_map[k], v) for k, v in _SCORE_MAP.items() if k in header_map]

    _append_marks_out_of_columns(header_map, score_mapping)

    if not score_mapping and assessment_field:
        gen_key = next((k for k in _GENERIC_SCORE_KEYS if k in header_map), None)
        if gen_key:
            score_mapping = [(header_map[gen_key], assessment_field)]

    # One generic "marks/score" column and no CT1/2/3 headers: default to Class Test 1 (ct_score1).
    if not score_mapping and not assessment_field:
        gen_key = next((k for k in _GENERIC_SCORE_KEYS if k in header_map), None)
        if gen_key:
            score_mapping = [(header_map[gen_key], "ct_score1")]

    if not score_mapping:
        return ParseResult(
            errors=[
                ParseError(
                    row=0,
                    reason=(
                        "No score columns found. Add CT1, CT2, CT3, and/or Lab headers, "
                        "or a single Marks/Score column (stored as Class Test 1)."
                    ),
                )
            ]
        )

    records: list[ParsedMarkRow] = []
    errors: list[ParseError] = []
    seen_norm: set[str] = set()

    for i, row_data in enumerate(rows):
        row_num = i + 2
        roll_value = str(row_data.get(roll_header, "")).strip()
        roll_norm = _normalize_roll(roll_value)

        if not roll_norm:
            if all(str(row_data.get(h, "") or "").strip() == "" for h in headers):
                continue
            errors.append(ParseError(row=row_num, reason="Missing roll number"))
            continue

        if roll_norm in seen_norm:
            errors.append(
                ParseError(row=row_num, roll_number=roll_value, reason="Duplicate roll number (later entry used)")
            )
            prev = next((j for j, r in enumerate(records) if _normalize_roll(r.roll_number) == roll_norm), None)
            if prev is not None:
                records.pop(prev)
        seen_norm.add(roll_norm)

        record = ParsedMarkRow(row=row_num, roll_number=roll_value)
        has_valid = False

        for orig_header, field_name in score_mapping:
            raw = row_data.get(orig_header)
            if raw is None or raw == "":
                continue
            try:
                num = float(raw)
            except (ValueError, TypeError):
                errors.append(
                    ParseError(
                        row=row_num,
                        roll_number=roll_value,
                        reason=f'Invalid value for {orig_header}: "{raw}"',
                    )
                )
                continue
            if num < 0:
                errors.append(
                    ParseError(
                        row=row_num,
                        roll_number=roll_value,
                        reason=f"Negative value for {orig_header}: {num}",
                    )
                )
                continue
            setattr(record, field_name, num)
            has_valid = True

        if has_valid:
            records.append(record)
        else:
            errors.append(
                ParseError(row=row_num, roll_number=roll_value, reason="No valid score values found")
            )

    return ParseResult(records=records, errors=errors)


def parse_marks_spreadsheet(data: bytes, filename: str, assessment_field: str | None) -> ParseResult:
    loaded = _load_tabular_rows(data, filename)
    if isinstance(loaded, ParseError):
        return ParseResult(errors=[loaded])
    return _parse_rows(loaded, assessment_field)


def parse_spreadsheet(data: bytes, filename: str = "") -> ParseResult:
    """Backward-compatible alias: multi-column files without a single-column assessment hint."""
    return parse_marks_spreadsheet(data, filename, None)
